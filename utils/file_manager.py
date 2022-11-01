from numpy import true_divide
from openpyxl import load_workbook
import pandas as pd

def insert_row(file_name, row_data):
    try:
        wb = load_workbook("files/{file_name}.xlsx".format(file_name=file_name))
        ws = wb.worksheets[0]
        ws.append(row_data)
        wb.save("files/{file_name}.xlsx".format(file_name=file_name))
        return True
    except:
        return False

def get_table(file_name):
    games = []
    wb = load_workbook("files/{file_name}.xlsx".format(file_name=file_name))
    ws = wb.worksheets[0]
    all_rows = list(ws.rows)

    for row in all_rows[1:]:
        curr_game = []
        for cell in row:
            curr_game.append(cell.value)
        games.append(curr_game)

    return games


    # read_file  = pd.read_excel("files/{file_name}.xlsx".format(file_name=file_name))
    # read_file.to_csv ("files/{file_name}.xlsx".format(file_name=file_name),index = None,header=True) 
    # # enter code here
    # df = pd.DataFrame(pd.read_csv("files/{file_name}.xlsx".format(file_name=file_name))) 
    # print(df)

    # for entry, data_boundary in ws.tables.items():
    #     # parse the data within the ref boundary
    #     data = ws[data_boundary]
        
    #     ### extract the data ###
    #     # the inner list comprehension gets the values for each cell in the table
    #     content = [[cell.value for cell in ent] 
    #             for ent in data]
        
    #     header = content[0]
        
    #     #the contents ... excluding the header
    #     rest = content[1:]
        
    #     #create dataframe with the column names
    #     #and pair table name with dataframe
    #     df = pd.DataFrame(rest, columns = header)
    #     mapping[entry] = df
    
    # print(mapping)


def search_in_table(file_name, user_name, user_password):
    wb = load_workbook("files/{file_name}.xlsx".format(file_name=file_name))
    ws = wb.worksheets[0]
    all_rows = list(ws.rows)

    for row in all_rows[1:]:
        if row[0].value == user_name:
            if row[1].value == user_password:
                return 'User connected'
            elif row[1].value != user_password:
                return 'Wrong password'
        else: continue

    return "User not exist!"
    